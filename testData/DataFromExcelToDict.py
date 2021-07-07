import openpyxl


def dataExcel(excel_file_path, test_case_name):
    excel_file = openpyxl.load_workbook(excel_file_path)
    active_sheet = excel_file.active
    Data_Dict = {}
    for _row in range(1, active_sheet.max_row + 1):
        if active_sheet.cell(row=_row, column=1).value == test_case_name:
            for _column in range(2, active_sheet.max_column + 1):
                Data_Dict[active_sheet.cell(row=1, column=_column).value] = active_sheet.cell(row=_row,
                                                                                              column=_column).value

    return Data_Dict
