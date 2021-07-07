# This file contains the data for home page elements,
# we define the data for different webpages in different data files under testData package
# If we want to access the methods of any class without creating its object,
#           then create the method as static by using decorator @staticmethod above the method definition
#           And also there is no need of self keyword, bcoz we are not creating object of class.
import openpyxl


class HomePageData:
    # def __init__(self, excel_file_path, test_case_name):
    #     self.test_case_name = test_case_name
    #     self.excel_file_path = excel_file_path

    @staticmethod
    def getDataExcel(excel_file_path, test_case_name):
        excel_file = openpyxl.load_workbook(excel_file_path)
        active_sheet = excel_file.active
        Data_Dict = {}
        for _row in range(1, active_sheet.max_row + 1):
            if active_sheet.cell(row=_row, column=1).value == test_case_name:
                for _column in range(2, active_sheet.max_column + 1):
                    Data_Dict[active_sheet.cell(row=1, column=_column).value] = active_sheet.cell(row=_row,
                                                                                                  column=_column).value

        return [Data_Dict]
